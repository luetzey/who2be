"""Unit-Tests (DB-los) fuer die Export-Renderer in `services/wa_render`.

Getestet wird die AUSGABE-Seite der WorkArea-Exporte (ADR-0047/0049) — reine
Funktionen, deshalb ohne Datenbank und ohne `integration`-Marker:

- `render_table_xlsx`: der Formel-Guard. Der Test liest die erzeugte Datei mit
  openpyxl ZURUECK und pinnt dabei den TYP der Zelle, nicht nur unseren String
  — inklusive Gegenprobe, dass dieselbe Eingabe OHNE Guard als Formelzelle
  ankaeme (Security-Review L5, OWASP CSV/Formula Injection).
- `render_artifact_export_markdown`: Frontmatter-Quoting und ausgelassene
  None-Felder.
- `render_artifact_export_html`: Markdown-Inhalt und Titel kommen escaped an,
  nicht als Tag (`html: False` im MarkdownIt + `html.escape`).
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

# openpyxl liefert kein `py.typed` — dieselbe eng begrenzte Ausnahme wie an der
# Import-Zeile in `services/wa_render`. (Faellt weg, sobald `types-openpyxl` in
# der dev-Gruppe steht; dann meldet mypy den ignore als unbenutzt.)
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from who2be_api.services.wa_render import (
    csv_cell,
    render_artifact_export_html,
    render_artifact_export_markdown,
    render_table_xlsx,
)
from who2be_models.workarea import OccurredPrecision, Sensitivity

_FORMULA = '=HYPERLINK("http://evil","x")'


def test_xlsx_formel_zelle_bleibt_string_und_pinnt_openpyxl_verhalten() -> None:
    """Formel-Praefix wird entschaerft — belegt am Zell-TYP der Datei."""
    payload = render_table_xlsx("Ergebnis", ["cmd"], [[_FORMULA]])

    sheet = load_workbook(BytesIO(payload)).worksheets[0]
    cell = sheet.cell(row=2, column=1)
    assert cell.value == f"'{_FORMULA}"
    assert cell.data_type == "s"
    assert cell.data_type != "f"

    # Gegenprobe: OHNE Guard leitet openpyxl aus demselben Praefix eine
    # Formelzelle ab — erst das macht den Test zu einem Beleg fuer die
    # Entschaerfung statt fuer eine beliebige String-Konvention.
    raw_workbook = Workbook()
    raw_workbook.worksheets[0]["A1"] = "=SUM(1)"
    raw_buffer = BytesIO()
    raw_workbook.save(raw_buffer)
    raw_cell = load_workbook(BytesIO(raw_buffer.getvalue())).worksheets[0]["A1"]
    assert raw_cell.data_type == "f"


def test_xlsx_native_typen_bleiben_zahl_bool_und_leer() -> None:
    """int/float/bool/None kommen nativ an — Zahlen bleiben rechenbar."""
    payload = render_table_xlsx("Werte", ["i", "f", "b", "n", "neg"], [[42, 3.5, True, None, -3.2]])

    sheet = load_workbook(BytesIO(payload)).worksheets[0]
    assert sheet.cell(row=1, column=1).value == "i"
    assert sheet.cell(row=2, column=1).value == 42
    assert sheet.cell(row=2, column=1).data_type == "n"
    assert sheet.cell(row=2, column=2).value == 3.5
    assert sheet.cell(row=2, column=3).value is True
    assert sheet.cell(row=2, column=3).data_type == "b"
    assert sheet.cell(row=2, column=4).value is None
    # Negative Zahl aus einer NUMERIC-Spalte ist eine ZAHL, keine Formel.
    assert sheet.cell(row=2, column=5).value == -3.2
    assert sheet.cell(row=2, column=5).data_type == "n"


def test_xlsx_sheet_name_wird_bereinigt_und_gekappt() -> None:
    """Excel-verbotene Zeichen und die 31-Zeichen-Grenze halten die Datei lesbar."""
    payload = render_table_xlsx("Bericht/2026 [Q3] mit sehr langem Namen", ["a"], [["x"]])

    title = load_workbook(BytesIO(payload)).sheetnames[0]
    assert len(title) <= 31
    assert not set(title) & set("[]:*?/\\")
    assert title == "Bericht_2026 _Q3_ mit sehr lang"


def test_markdown_export_quotet_frontmatter_und_laesst_none_weg() -> None:
    """Frontmatter-Quoting nach dem `entity_export_service`-Muster."""
    doc = render_artifact_export_markdown(
        title='Titel: "mit: Doppelpunkt" und Umlauten äöü',
        markdown="# Ueberschrift\n\nInhalt.",
        occurred_at=datetime(2026, 8, 19, 10, 30, tzinfo=UTC),
        occurred_precision=OccurredPrecision.minute,
        sensitivity=Sensitivity.general,
        source_system=None,
        source_url=None,
        exported_at=datetime(2026, 8, 19, 12, 0, tzinfo=UTC),
    )

    head, _, body = doc.partition("---\n\n")
    assert head.startswith("---\n")
    assert 'title: "Titel: \\"mit: Doppelpunkt\\" und Umlauten äöü"' in head
    assert 'occurred_at: "2026-08-19T10:30:00+00:00"' in head
    assert 'occurred_precision: "minute"' in head
    assert 'sensitivity: "general"' in head
    assert 'exported_at: "2026-08-19T12:00:00+00:00"' in head
    assert "source_system" not in head
    assert "source_url" not in head
    assert body == "# Ueberschrift\n\nInhalt."


def test_markdown_export_nimmt_gesetzte_quellfelder_auf() -> None:
    """Gesetzte Quellfelder landen gequotet im Frontmatter."""
    doc = render_artifact_export_markdown(
        title="Bericht",
        markdown="Text.",
        occurred_at=None,
        occurred_precision=OccurredPrecision.unknown,
        sensitivity=Sensitivity.sensitive,
        source_system="notion",
        source_url='https://example.test/a?b="c"',
        exported_at=datetime(2026, 8, 19, 12, 0, tzinfo=UTC),
    )

    assert 'source_system: "notion"' in doc
    assert 'source_url: "https://example.test/a?b=\\"c\\""' in doc
    assert "occurred_at" not in doc


def test_html_export_escapet_markdown_und_titel() -> None:
    """Rohes HTML im Markdown und im Titel wird Text, nicht Markup."""
    doc = render_artifact_export_html(
        title="<b>Titel</b>",
        markdown="Hallo <script>alert(1)</script>",
    )

    assert doc.startswith("<!doctype html>")
    assert '<meta charset="utf-8">' in doc
    assert "<script>" not in doc
    assert "&lt;script&gt;alert(1)&lt;/script&gt;" in doc
    # Einmal im <title>, einmal in der <h1>.
    assert doc.count("&lt;b&gt;Titel&lt;/b&gt;") == 2


def test_xlsx_steuerzeichen_werden_entfernt_statt_500() -> None:
    """M-2: XML-illegale Steuerzeichen aus Bestandsdaten duerfen den Export
    nicht abbrechen — openpyxl wuerde sonst `IllegalCharacterError` werfen."""
    data = render_table_xlsx("t", ["c"], [["a\x01b\x00c"]])
    sheet = load_workbook(BytesIO(data)).worksheets[0]
    assert sheet["A2"].value == "abc"


def test_frontmatter_neutralisiert_zeilenumbruch_in_quellfeldern() -> None:
    """M-3: ein roher Zeilenumbruch in `source_system` wuerde die
    Frontmatter-Struktur aufbrechen (Parser schneiden am ersten `---`)."""
    exported = render_artifact_export_markdown(
        title="T",
        markdown="Body",
        occurred_at=None,
        occurred_precision=None,
        sensitivity=None,
        source_system="evil\n---\n\n# Gefaelscht",
        source_url=None,
        exported_at=datetime(2026, 8, 19, 12, 0, tzinfo=UTC),
    )
    frontmatter = exported.split("---")[1]
    assert "\n# Gefaelscht" not in frontmatter
    assert 'source_system: "evil --- # Gefaelscht"' in exported


def test_csv_guard_greift_auch_hinter_fuehrendem_whitespace() -> None:
    """L-2: Google Sheets trimmt beim Import fuehrenden Whitespace und wertet
    dann aus — der Guard muss auf der getrimmten Kopie pruefen."""
    for value in (" =1+1", "﻿=1+1", "\xa0=1+1", "\n=1+1"):
        assert str(csv_cell(value)).startswith("'"), repr(value)
    # Legitimer Text mit fuehrendem Leerzeichen bleibt unangetastet.
    assert csv_cell(" normaler Text") == " normaler Text"


def test_html_export_traegt_meta_csp_und_no_referrer() -> None:
    """L-1: der Export wird aus `file://` geoeffnet — nur eine Meta-CSP im
    Dokument selbst verhindert, dass Bild-Ziele aus Fremdquellen laden."""
    doc = render_artifact_export_html(title="T", markdown="![x](https://evil.example/t.png)")
    assert 'http-equiv="Content-Security-Policy"' in doc
    assert "default-src 'none'" in doc
    assert '<meta name="referrer" content="no-referrer">' in doc
